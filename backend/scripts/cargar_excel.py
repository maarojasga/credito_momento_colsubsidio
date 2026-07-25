"""Carga afiliados desde un Excel y corre el pipeline (reemplaza los datos demo).

El Excel aporta los AFILIADOS (lo que un piloto sin buró sí tiene de PILA /
afiliación). El historial de eventos de servicio de cada afiliado se SIMULA con
el proceso generador (`timing/sequences.py`) para ilustrar el timing, porque en
un piloto real ese historial aún no existe. El modelo de hazard se entrena sobre
una población de referencia (ver hazard.ajustar_hazard).

Privacidad: solo se guarda el `subject_id` = hash del documento; nombre, correo y
documento del Excel nunca se almacenan (principio "hash del documento, nunca el
documento").

Columnas esperadas (encabezados en la primera fila, sin importar mayúsculas):
    documento, categoria, sexo, edad, ingreso_mensual,
    antiguedad_empleo_meses, contrato_indefinido, estrato, localidad, num_hijos
(nombre y correo son opcionales y se ignoran al almacenar).

Uso:
    python scripts/cargar_excel.py afiliados.xlsx --db data/synthetic/momento.duckdb
"""

from __future__ import annotations

import argparse
import hashlib
import time
from datetime import date

from openpyxl import load_workbook

from momento.pipeline import ejecutar_pipeline
from momento.storage import connect, load_synthetic
from momento.timing.params import SMMLV
from momento.timing.sequences import SyntheticDataset, generar_eventos


def _bool(v) -> bool:
    return str(v).strip().lower() in ("si", "sí", "true", "1", "x", "indefinido", "verdadero", "y")


def _int(v, defecto: int = 0) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return defecto


def leer_afiliados(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []
    encabezados = [str(h).strip().lower() if h is not None else "" for h in filas[0]]

    subjects: list[dict] = []
    for fila in filas[1:]:
        if fila is None or all(c is None for c in fila):
            continue
        d = dict(zip(encabezados, fila))
        documento = str(d.get("documento") or "").strip()
        if not documento:
            continue
        sid = "sub_" + hashlib.sha256(documento.encode()).hexdigest()[:16]
        ingreso_mensual = _int(d.get("ingreso_mensual"), 0)
        subjects.append({
            "subject_id": sid,
            "categoria": str(d.get("categoria") or "A").strip().upper()[:1],
            "edad": _int(d.get("edad"), 40),
            "sexo": str(d.get("sexo") or "M").strip().upper()[:1],
            "ingreso_smmlv": round(ingreso_mensual / SMMLV, 2) if ingreso_mensual else 1.0,
            "ingreso_mensual": ingreso_mensual,
            "antiguedad_empleo_meses": _int(d.get("antiguedad_empleo_meses"), 0),
            "contrato_indefinido": _bool(d.get("contrato_indefinido")),
            "num_hijos": _int(d.get("num_hijos"), 0),
            "localidad": str(d.get("localidad") or "").strip(),
            "estrato": _int(d.get("estrato"), 3),
        })
    return subjects


def main() -> None:
    ap = argparse.ArgumentParser(description="Carga afiliados desde Excel y corre el pipeline")
    ap.add_argument("excel", help="ruta al .xlsx de afiliados")
    ap.add_argument("--db", default="data/synthetic/momento.duckdb")
    ap.add_argument("--as-of", default="2026-07-01")
    args = ap.parse_args()

    t0 = time.perf_counter()
    subjects = leer_afiliados(args.excel)
    if not subjects:
        raise SystemExit("El Excel no tiene afiliados válidos (revisa la columna 'documento').")

    eventos, person_period = generar_eventos(subjects, seed=42)
    ds = SyntheticDataset(subjects=subjects, eventos=eventos, person_period=person_period)

    con = connect(args.db)
    load_synthetic(con, ds)                 # reemplaza los datos demo
    r = ejecutar_pipeline(con, date.fromisoformat(args.as_of))
    con.close()

    print(f"Cargados {len(subjects)} afiliados del Excel → {args.db}")
    print(f"{r['ofertas']} ofertas · {r['no_elegibles']} no elegibles · {time.perf_counter()-t0:.1f}s")


if __name__ == "__main__":
    main()
