"""Genera la plantilla `afiliados.xlsx` con datos de ejemplo.

Los datos son ficticios y variados para ejercitar los cuatro productos y los
casos de rechazo (antigüedad insuficiente, contrato temporal corto). Reemplaza
las filas con tus afiliados reales manteniendo los encabezados.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

COLUMNAS = [
    "documento", "nombre", "correo", "categoria", "sexo", "edad",
    "ingreso_mensual", "antiguedad_empleo_meses", "contrato_indefinido",
    "estrato", "localidad", "num_hijos",
]

# Filas ficticias (nombre/correo solo para referencia del operador; no se almacenan).
EJEMPLOS = [
    [1001, "Ana Gómez", "ana@ejemplo.co", "B", "F", 34, 3200000, 40, "SI", 3, "Suba", 2],
    [1002, "Carlos Ruiz", "carlos@ejemplo.co", "A", "M", 28, 1500000, 8, "NO", 2, "Kennedy", 0],
    [1003, "María López", "maria@ejemplo.co", "A", "F", 45, 1300000, 60, "SI", 2, "Bosa", 3],
    [1004, "Jorge Pérez", "jorge@ejemplo.co", "D", "M", 50, 11000000, 120, "SI", 5, "Chapinero", 1],
    [1005, "Laura Díaz", "laura@ejemplo.co", "C", "F", 38, 5200000, 24, "SI", 4, "Usaquén", 2],
    [1006, "Pedro Sánchez", "pedro@ejemplo.co", "B", "M", 41, 3800000, 4, "NO", 3, "Engativá", 2],
    [1007, "Sofía Torres", "sofia@ejemplo.co", "A", "F", 25, 1400000, 1, "NO", 1, "Ciudad Bolívar", 0],
    [1008, "Andrés Castro", "andres@ejemplo.co", "C", "M", 55, 6500000, 90, "SI", 4, "Teusaquillo", 0],
    [1009, "Diana Moreno", "diana@ejemplo.co", "B", "F", 36, 3000000, 18, "SI", 3, "Fontibón", 1],
    [1010, "Felipe Rojas", "felipe@ejemplo.co", "D", "M", 48, 9000000, 72, "SI", 6, "Usaquén", 2],
]

AZUL = "0067B1"
AMARILLO = "FFD000"


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "afiliados"

    ws.append(COLUMNAS)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(horizontal="center")

    for fila in EJEMPLOS:
        ws.append(fila)

    anchos = [11, 16, 20, 10, 6, 6, 15, 22, 20, 8, 16, 10]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    salida = Path(__file__).resolve().parents[1] / "afiliados.xlsx"
    wb.save(salida)
    print(f"Plantilla escrita en {salida} ({len(EJEMPLOS)} afiliados de ejemplo)")


if __name__ == "__main__":
    main()
