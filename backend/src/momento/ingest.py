"""Ingesta de afiliados desde Excel y ejecución del pipeline.

Reutilizable desde el script `scripts/cargar_excel.py` (CLI) y desde el endpoint
de carga de la API (subida manual por la UI).

Privacidad: solo se guarda `subject_id` = hash del documento; nombre, correo y
documento del Excel nunca se almacenan.
"""

from __future__ import annotations

import hashlib
from datetime import date

from openpyxl import load_workbook

from momento.pipeline import ejecutar_pipeline
from momento.storage import connect, load_synthetic
from momento.timing.params import SMMLV
from momento.timing.sequences import SyntheticDataset, generar_eventos

# Columnas esperadas (encabezados en la primera fila, sin importar mayúsculas).
COLUMNAS = [
    "documento", "nombre", "correo", "categoria", "sexo", "edad",
    "ingreso_mensual", "antiguedad_empleo_meses", "contrato_indefinido",
    "estrato", "localidad", "num_hijos",
]


def _bool(v) -> bool:
    return str(v).strip().lower() in ("si", "sí", "true", "1", "x", "indefinido", "verdadero", "y")


def _int(v, defecto: int = 0) -> int:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return defecto


def leer_afiliados(path: str) -> list[dict]:
    """Lee el Excel y devuelve la lista de sujetos (sin PII, documento hasheado)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []
    encabezados = [str(h).strip().lower() if h is not None else "" for h in filas[0]]

    subjects: list[dict] = []
    for fila in filas[1:]:
        if fila is None or all(c is None for c in fila):
            continue
        d = dict(zip(encabezados, fila))
        documento = str(d.get("documento") or "").strip()
        if not documento:
            continue
        sid = "sub_" + hashlib.sha256(documento.encode()).hexdigest()[:16]
        ingreso_mensual = _int(d.get("ingreso_mensual"), 0)
        subjects.append({
            "subject_id": sid,
            "categoria": str(d.get("categoria") or "A").strip().upper()[:1],
            "edad": _int(d.get("edad"), 40),
            "sexo": str(d.get("sexo") or "M").strip().upper()[:1],
            "ingreso_smmlv": round(ingreso_mensual / SMMLV, 2) if ingreso_mensual else 1.0,
            "ingreso_mensual": ingreso_mensual,
            "antiguedad_empleo_meses": _int(d.get("antiguedad_empleo_meses"), 0),
            "contrato_indefinido": _bool(d.get("contrato_indefinido")),
            "num_hijos": _int(d.get("num_hijos"), 0),
            "localidad": str(d.get("localidad") or "").strip(),
            "estrato": _int(d.get("estrato"), 3),
        })
    return subjects


def ingestar_excel(xlsx_path: str, db_path: str, as_of: date) -> dict:
    """Lee el Excel, reemplaza los datos y corre el pipeline. Devuelve el resumen.

    El historial de eventos de cada afiliado se simula para ilustrar el timing
    (en un piloto sin buró ese historial aún no existe).
    """
    subjects = leer_afiliados(xlsx_path)
    if not subjects:
        raise ValueError("El Excel no tiene afiliados válidos (revisa la columna 'documento').")

    eventos, person_period = generar_eventos(subjects, seed=42)
    ds = SyntheticDataset(subjects=subjects, eventos=eventos, person_period=person_period)

    con = connect(db_path)
    load_synthetic(con, ds)              # reemplaza los datos previos
    resumen = ejecutar_pipeline(con, as_of)
    con.close()
    resumen["afiliados"] = len(subjects)
    return resumen
